# -------------------------------
# IMPORTACIÓN DE LIBRERÍAS
# -------------------------------
import os          # Para trabajar con archivos y carpetas
import re          # Expresiones regulares
import requests    # Realizar solicitudes HTTP (para el OCR)
import zipfile     # Manejar archivos ZIP
import patoolib    # Manejar archivos RAR
import shutil      # Funciones de copiado/movido/eliminación
from openpyxl import Workbook  # Crear archivos Excel

# -------------------------------
# CONSTANTES DE CONFIGURACIÓN
# -------------------------------
API_KEY = "K87845793988957"        # Clave del servicio OCR
CARPETA_PDF = "pdfs"               # Carpeta donde están los ZIP/RAR a procesar
CARPETA_EXTRAIDA = "extraido"      # Carpeta donde se extraerán los PDF encontrados

# Creamos las carpetas necesarias si no existen
os.makedirs(CARPETA_PDF, exist_ok=True)
os.makedirs(CARPETA_EXTRAIDA, exist_ok=True)

# -------------------------------------------------------
# Función que descomprime archivos ZIP o RAR buscando PDF
# -------------------------------------------------------
def decomprimir_archivo(ruta):
    # Si el archivo es ZIP
    if ruta.lower().endswith(".zip"):
        with zipfile.ZipFile(ruta, 'r') as zip_ref:
            # Recorremos cada archivo que contiene el ZIP
            for nombre_archivo in zip_ref.namelist():
                # Sólo extraemos los PDF que contengan "reporte" en su nombre
                if nombre_archivo.lower().endswith(".pdf") and "reporte" in nombre_archivo.lower():
                    # Armamos ruta destino
                    destino = os.path.join(CARPETA_EXTRAIDA, os.path.basename(nombre_archivo))

                    # Si ya existe un PDF con ese nombre, agregamos _1, _2, _3… para no sobrescribir
                    if os.path.exists(destino):
                        base, ext = os.path.splitext(os.path.basename(nombre_archivo))
                        contador = 1
                        nuevo = f"{base}_{contador}{ext}"
                        while os.path.exists(os.path.join(CARPETA_EXTRAIDA, nuevo)):
                            contador += 1
                            nuevo = f"{base}_{contador}{ext}"
                        destino = os.path.join(CARPETA_EXTRAIDA, nuevo)

                    # Extraemos el PDF del ZIP manualmente
                    with open(destino, 'wb') as f_out:
                        f_out.write(zip_ref.read(nombre_archivo))
                    print(f"ZIP: Extraído como {os.path.basename(destino)}")

    # Si el archivo es RAR
    elif ruta.lower().endswith(".rar"):
        carpeta_temp = "temp_rar"
        os.makedirs(carpeta_temp, exist_ok=True)

        try:
            # Extraemos todo el RAR a una carpeta temporal
            patoolib.extract_archive(ruta, outdir=carpeta_temp, verbosity=-1)
        except Exception as e:
            print(f"Error al extraer RAR: {ruta} -> {e}")
            return False

        # Recorremos los archivos extraídos
        for archivo in os.listdir(carpeta_temp):
            # Si es un PDF con 'reporte' en el nombre
            if archivo.lower().endswith(".pdf") and "reporte" in archivo.lower():
                origen = os.path.join(carpeta_temp, archivo)
                destino = os.path.join(CARPETA_EXTRAIDA, archivo)

                # Evitar sobrescritura creando nombre incremental
                if os.path.exists(destino):
                    base, ext = os.path.splitext(archivo)
                    contador = 1
                    nuevo = f"{base}_{contador}{ext}"
                    while os.path.exists(os.path.join(CARPETA_EXTRAIDA, nuevo)):
                        contador += 1
                        nuevo = f"{base}_{contador}{ext}"
                    destino = os.path.join(CARPETA_EXTRAIDA, nuevo)

                shutil.move(origen, destino)
                print(f"RAR: Extraído como {os.path.basename(destino)}")

        # Eliminamos la carpeta temporal
        shutil.rmtree(carpeta_temp)

    else:
        print(f"Formato no reconocido: {os.path.basename(ruta)}")
        return False

    return True

# -------------------------------------------------------
# Buscar PDF que contengan 'reporte' dentro de "extraido/"
# -------------------------------------------------------
def buscar_pdfs_reporte():
    return [
        os.path.join(CARPETA_EXTRAIDA, f)
        for f in os.listdir(CARPETA_EXTRAIDA)
        if f.lower().endswith(".pdf") and "reporte" in f.lower()
    ]

# -------------------------------------------------------
# Enviar PDF al servicio OCR.space y obtener texto reconocido
# -------------------------------------------------------
def extraer_texto_ocr(pdf_path):
    url_api = "https://api.ocr.space/parse/image"
    payload = {
        'apikey': API_KEY,
        'language': 'spa',        # idioma español
        'isOverlayRequired': False,
        'OCREngine': 2,
        'isTable': True           # mejor detección si hay tablas
    }

    with open(pdf_path, 'rb') as f:
        r = requests.post(url_api, files={'file': f}, data=payload)

    try:
        return r.json()["ParsedResults"][0]["ParsedText"]
    except:
        print("Error procesando OCR en:", pdf_path)
        return ""

# -------------------------------------------------------
# Buscar palabra "Resultado .... Adjudicado" en el texto
# -------------------------------------------------------
def buscar_resultado(texto):
    texto = texto.lower()
    for match in re.finditer(r"resultado.{0,9}?(adjudicado)", texto):
        return match.group(1).capitalize()
    return None



# -------------------------------------------------------
# Extraer monto adjudicado
# -------------------------------------------------------
def extraer_monto(texto):
    lineas = texto.splitlines()
    for i, linea in enumerate(lineas):
        if "monto adjudicado" in linea.lower():
            if i + 1 < len(lineas):
                monto_linea = lineas[i + 1].strip()
                # Limpiar caracteres comunes de OCR
                monto_linea = monto_linea.replace(',', '').replace('o', '0')
                # Buscar patrón de número (opcionalmente con símbolo)
                match = re.search(r"(S\/\.|US\$)?\s?(\d+(?:\.\d{1,2})?)", monto_linea)
                if match:
                    simbolo = match.group(1) or ""
                    valor = match.group(2)
                    return simbolo + valor
                else:
                    return monto_linea
    return "Monto no encontrado"

# ***********************************************************
#               BLOQUE PRINCIPAL DE PROCESAMIENTO
# ***********************************************************
def procesar_excel(excel_path, fila_excel):
    resultados = []

# 1) Recorrer todos los archivos ZIP o RAR dentro de "pdfs/"
    for archivo_comprimido in os.listdir(CARPETA_PDF):
        ruta = os.path.join(CARPETA_PDF, archivo_comprimido)
        if ruta.lower().endswith(".zip") or ruta.lower().endswith(".rar"):
           print(f"\nProcesando archivo comprimido: {archivo_comprimido}")
        decomprimir_archivo(ruta)

# 2) Buscar los PDF extraídos
    pdfs_reporte = buscar_pdfs_reporte()

    if not pdfs_reporte:
       print("No se encontró PDF con 'reporte' en el nombre.")
       resultado, monto = "Desierto", ""
    else:
       # 3) Para cada PDF -> ejecutar OCR, extraer info, y guardar resultado
       for ruta_pdf in pdfs_reporte:
        print(f"Evaluando: {os.path.basename(ruta_pdf)}")

        texto = extraer_texto_ocr(ruta_pdf)
        resultado = buscar_resultado(texto)

        if resultado == "Adjudicado":
            monto = extraer_monto(texto)
            print(f"Resultado: {resultado} | Monto: {monto}")
            
        else:
            print("No se encontró resultado 'Adjudicado'.")
            resultado, monto = "Desierto", ""

# 4) Si encontramos resultados válidos, los grabamos a Excel
    if resultados:
        wb = Workbook(excel_path)
        ws = wb.active
        ws.title = "Resultados"
     # Columna C = Resultado | Columna D = Monto
        ws.cell(row=fila_excel, column=33, value=resultado)
        ws.cell(row=fila_excel, column=34, value=monto)
    
        wb.save(excel_path)
        print(f"Fila {fila_excel} actualizada en {excel_path}: {resultado}, {monto}")
    else:
        print("\nNo se encontraron resultados válidos para guardar.")

procesar_excel()


##BUENO MODIFICANDO CODIGO PEg